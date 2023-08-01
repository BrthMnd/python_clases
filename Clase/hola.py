# # import numpy as np
# import random as rd
# import numpy as np

# arr = np.array([1, 2, 3, 4, 5])

# print(arr)

# # list = [1,2,3,4,5,6,7,8,9,10]
# list = []
# for i in range(20):
#     num = rd.randint(0,99)
#     list.append(num)
# list.sort()
# listanueva = [str(i)for i in list]
# cadena = "-".join(listanueva)


# # print(listanueva)
# print(cadena)


import openpyxl as ox
from faker import Faker
import random as rd


fake = Faker()
print(fake.name())

def CreateExcelFile():
    libro = ox.Workbook()
    hoja = libro.active
    hoja.title = "datos bb"
    hoja.append(["Nombre", "Edad", "Direccion"])
    for persona in range(1000000):
        hoja.append([fake.name(),rd.randint(1,30),fake.address()])
    libro.save("hola.xlsx")
    return 
CreateExcelFile()

def ReadExcelFile():
    data = []
    libro = ox.load_workbook("hola.xlsx")
    hoja = libro["datos bb"]
    for fila in hoja.iter_rows(min_row=2,values_only=True):
        nombre, edad, direccion = fila
        data.append({"nombre": nombre, "Edad": edad, "direccion": direccion})
    libro.close()
    print(data)
        
ReadExcelFile()